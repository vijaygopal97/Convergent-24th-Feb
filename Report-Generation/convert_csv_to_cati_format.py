#!/usr/bin/env python3
"""
CATI Respondent CSV to Excel Converter

This script converts CSV files with respondent data to the CATI Respondent Template format.

Usage:
    python convert_csv_to_cati_format.py <input_csv> <output_excel> [master_data_excel]

Arguments:
    input_csv: Path to the input CSV file with respondent data
    output_excel: Path to the output Excel file
    master_data_excel: (Optional) Path to the master data Excel file containing AC mappings
                      Defaults to 'West_Bengal_Raw_Data_12th_Nov_2025.xlsx'
"""

import pandas as pd
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


def load_ac_mapping(master_data_path):
    """
    Load AC code to AC name mapping from master data file.
    
    Args:
        master_data_path: Path to the master data Excel file
        
    Returns:
        dict: Dictionary mapping AC codes to AC names
    """
    print(f"Loading AC mapping from: {master_data_path}")
    
    # Read the master data (skip the header row)
    master_data = pd.read_excel(master_data_path, skiprows=1)
    
    # Extract unique AC mappings
    ac_mapping_df = master_data[['ac_code', 'ac_name']].drop_duplicates()
    
    # Convert to dictionary
    ac_mapping = dict(zip(ac_mapping_df['ac_code'], ac_mapping_df['ac_name']))
    
    print(f"Loaded {len(ac_mapping)} AC mappings")
    return ac_mapping


def convert_csv_to_cati_format(input_csv, output_excel, ac_name, master_data_path=None):
    """
    Convert CSV file to CATI Respondent Template format.
    
    Args:
        input_csv: Path to input CSV file
        output_excel: Path to output Excel file
        ac_name: Assembly Constituency name to assign to all contacts
        master_data_path: (Optional) Path to master data Excel file (not currently used)
    """
    print(f"\n{'='*60}")
    print(f"Converting: {os.path.basename(input_csv)}")
    print(f"AC Name: {ac_name}")
    print(f"{'='*60}\n")
    
    # Load AC mapping (if master data provided, though not currently used for AC assignment)
    if master_data_path and os.path.exists(master_data_path):
        ac_mapping = load_ac_mapping(master_data_path)
    else:
        ac_mapping = {}
        print("Note: Master data file not provided or not found. AC name will be set directly.")
    
    # Read input CSV with error handling for malformed rows and encoding issues
    print(f"Reading input CSV: {input_csv}")
    # Try different encodings
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'utf-8-sig']
    df = None
    for encoding in encodings:
        try:
            df = pd.read_csv(input_csv, on_bad_lines='skip', engine='python', encoding=encoding)
            print(f"Successfully loaded with encoding: {encoding}")
            break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            print(f"Error with encoding {encoding}: {e}")
            continue
    
    if df is None:
        raise Exception("Could not read CSV file with any supported encoding")
    
    print(f"Loaded {len(df)} records")
    
    # Create output dataframe with template columns
    output_df = pd.DataFrame()
    
    # Map the columns according to requirements
    print("\nMapping columns...")
    
    # Name column <- FM_NAME_EN
    if 'FM_NAME_EN' in df.columns:
        output_df['Name'] = df['FM_NAME_EN']
        print(f"  - Name: Mapped from FM_NAME_EN ({df['FM_NAME_EN'].notna().sum()} records)")
    else:
        print("  - WARNING: FM_NAME_EN column not found!")
        output_df['Name'] = ''
    
    # Country Code <- leave empty
    output_df['Country Code'] = ''
    print(f"  - Country Code: Left empty")
    
    # Phone column <- MOBILE_NO (with +91 prefix removal)
    if 'MOBILE_NO' in df.columns:
        # Remove +91 prefix from phone numbers
        output_df['Phone'] = df['MOBILE_NO'].astype(str)
        # Replace 'nan' strings first
        output_df['Phone'] = output_df['Phone'].replace('nan', '')
        # Strip whitespace first
        output_df['Phone'] = output_df['Phone'].str.strip()
        # Remove +91 prefix (handles +91 at start, or whole string is +91)
        # This regex removes +91 from the beginning, optionally followed by space
        output_df['Phone'] = output_df['Phone'].str.replace(r'^\+91\s*', '', regex=True)
        # Also handle cases where +91 appears multiple times
        output_df['Phone'] = output_df['Phone'].str.replace(r'\+91', '', regex=True)
        # Strip whitespace again after removal
        output_df['Phone'] = output_df['Phone'].str.strip()
        print(f"  - Phone: Mapped from MOBILE_NO ({df['MOBILE_NO'].notna().sum()} records)")
        print(f"    * Removed +91 prefix from phone numbers")
    else:
        print("  - WARNING: MOBILE_NO column not found!")
        output_df['Phone'] = ''
    
    # Email column <- leave empty
    output_df['Email'] = ''
    
    # Address column <- leave empty
    output_df['Address'] = ''
    
    # City column <- leave empty
    output_df['City'] = ''
    
    # AC column <- Set to provided AC name for all rows
    output_df['AC'] = ac_name
    print(f"  - AC: Set to '{ac_name}' for all {len(output_df)} records")
    
    # Other columns <- leave empty
    output_df['PC'] = ''
    output_df['PS'] = ''
    
    print(f"  - Email, Address, City, PC, PS: Left empty")
    
    # Filter out records without phone numbers (including those that became empty after +91 removal)
    initial_count = len(output_df)
    # Convert phone to string and check for valid non-empty values
    phone_str = output_df['Phone'].astype(str)
    # Keep only records where phone is not null, not empty, not 'nan', and has actual digits
    valid_phone_mask = (
        output_df['Phone'].notna() & 
        (phone_str != '') & 
        (phone_str != 'nan') &
        (phone_str.str.strip() != '') &
        (phone_str.str.len() > 0)
    )
    output_df = output_df[valid_phone_mask]
    filtered_count = len(output_df)
    removed_count = initial_count - filtered_count
    
    print(f"\nFiltering records without phone numbers...")
    print(f"  - Initial records: {initial_count}")
    print(f"  - Records with valid phone numbers (after +91 removal): {filtered_count}")
    print(f"  - Records removed (no phone or empty after +91 removal): {removed_count}")
    
    # Save to Excel
    print(f"\nSaving output to: {output_excel}")
    output_df.to_excel(output_excel, index=False, engine='openpyxl')
    
    # Format the Excel file to match template style
    wb = load_workbook(output_excel)
    ws = wb.active
    
    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_excel)
    
    print(f"\n{'='*60}")
    print(f"Conversion completed successfully!")
    print(f"{'='*60}")
    print(f"\nOutput Summary:")
    print(f"  - Total records: {len(output_df)}")
    print(f"  - Output file: {output_excel}")
    print(f"  - File size: {os.path.getsize(output_excel) / 1024:.2f} KB")
    print()


def main():
    """Main function to handle command line arguments and execute conversion."""
    
    # Check arguments
    if len(sys.argv) < 4:
        print("Error: Insufficient arguments")
        print("\nUsage:")
        print("  python convert_csv_to_cati_format.py <input_csv> <output_excel> <ac_name> [master_data_excel]")
        print("\nExample:")
        print("  python convert_csv_to_cati_format.py ac171.csv output.xlsx \"Kharagpur\"")
        print("  python convert_csv_to_cati_format.py ac171.csv output.xlsx \"Kharagpur\" West_Bengal_Raw_Data.xlsx")
        sys.exit(1)
    
    input_csv = sys.argv[1]
    output_excel = sys.argv[2]
    ac_name = sys.argv[3]
    
    # Master data file is optional (4th argument)
    master_data_path = None
    if len(sys.argv) >= 5:
        master_data_path = sys.argv[4]
    
    # Validate input file exists
    if not os.path.exists(input_csv):
        print(f"Error: Input CSV file not found: {input_csv}")
        sys.exit(1)
    
    # Execute conversion
    try:
        convert_csv_to_cati_format(input_csv, output_excel, ac_name, master_data_path)
    except Exception as e:
        print(f"\nError during conversion: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()

